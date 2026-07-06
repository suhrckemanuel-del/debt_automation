from __future__ import annotations

import hashlib
import io
import zipfile
from datetime import datetime, timezone
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

TEXT_FORMAT = "@"
PERCENT_FORMAT = '0.00"%"'
AMOUNT_FORMAT = "#,##0.00"
REVIEW_SENTENCE = (
    "This is not a conclusion of compliance, breach, default, waiver "
    "satisfaction, or transaction permission."
)
PARITY_BANNER = (
    "Engine Decimal values (column B) are authoritative. Column C is an "
    "independent Excel recomputation in binary floating point, shown as a "
    "cross-check only."
)
CALCULATION_FOOTER = (
    "These are live Excel formulas recomputing the engine's arithmetic in "
    "IEEE-754 binary floating point. They are a cross-check, not the "
    "authoritative value - see the Parity sheet."
)
INPUTS_FOOTER = (
    "Every number above carries its exact document, locator, page, and "
    "quoted passage. Compare this sheet to the Parity sheet before relying "
    "on the Calculation sheet's arithmetic."
)
SUMMARY_PROVENANCE_CELL = "B20"

_REQUIRED_KEYS = {
    "model_id",
    "model_version",
    "status",
    "calculation_purpose",
    "evaluation_date",
    "test_date",
    "currency",
    "scenario",
    "source_inputs",
    "calculation_inputs",
    "formula",
    "outputs",
    "selected_threshold",
    "waiver_observation",
    "assumptions",
    "human_review_required",
    "review_note",
    "sources",
}

_PARITY_METRICS = (
    ("ltv_percent", "B2"),
    ("headroom_percentage_points", "B3"),
    ("maximum_debt_at_threshold", "B4"),
    ("debt_capacity_headroom", "B5"),
    ("minimum_valuation_at_threshold", "B6"),
)


class WorkbookGenerationAbstained(ValueError):
    """Raised when the calculation payload is not a calculated result."""

    def __init__(self, reason: str, missing_information: list[str]) -> None:
        super().__init__(reason)
        self.reason = reason
        self.missing_information = missing_information


class CitationRoles:
    def __init__(
        self,
        debt: dict[str, Any],
        valuation: dict[str, Any],
        threshold: dict[str, Any],
        definition: dict[str, Any],
        waiver: dict[str, Any] | None,
    ) -> None:
        self.debt = debt
        self.valuation = valuation
        self.threshold = threshold
        self.definition = definition
        self.waiver = waiver


def _status_label(status: str) -> str:
    return status.replace("_", " ").replace("selected ", "")


def _resolve_citation_roles(calculation: dict[str, Any]) -> CitationRoles:
    sources = list(calculation["sources"])

    def take_unique(document_id: str, locator: str) -> dict[str, Any]:
        matches = [
            source
            for source in sources
            if source["document_id"] == document_id
            and source["locator"] == locator
        ]
        if len(matches) != 1:
            raise WorkbookGenerationAbstained(
                "Source citation set does not match the expected "
                "calculation shape.",
                [],
            )
        sources.remove(matches[0])
        return matches[0]

    debt_input = calculation["source_inputs"]["debt_amount"]
    valuation_input = calculation["source_inputs"]["valuation_amount"]
    threshold_ref = calculation["selected_threshold"]
    debt = take_unique(debt_input["document_id"], debt_input["locator"])
    valuation = take_unique(
        valuation_input["document_id"], valuation_input["locator"]
    )
    threshold = take_unique(
        threshold_ref["document_id"], threshold_ref["locator"]
    )

    waiver: dict[str, Any] | None = None
    if calculation["waiver_observation"] is not None:
        waiver_matches = [
            source
            for source in sources
            if source["document_type"] == "waiver_letter"
        ]
        if len(waiver_matches) != 1:
            raise WorkbookGenerationAbstained(
                "Source citation set does not match the expected "
                "calculation shape.",
                [],
            )
        waiver = waiver_matches[0]
        sources.remove(waiver)

    if len(sources) != 1:
        raise WorkbookGenerationAbstained(
            "Source citation set does not match the expected "
            "calculation shape.",
            [],
        )
    return CitationRoles(
        debt=debt,
        valuation=valuation,
        threshold=threshold,
        definition=sources[0],
        waiver=waiver,
    )


def _review_banner_text(calculation: dict[str, Any]) -> str:
    return f"{calculation['review_note']} {REVIEW_SENTENCE}"


def _set_text(cell: Any, value: str) -> None:
    cell.value = value
    cell.number_format = TEXT_FORMAT


def _banner(
    sheet: Worksheet,
    cell_range: str,
    text: str,
    *,
    red_border: bool = False,
    shaded: bool = False,
) -> None:
    sheet.merge_cells(cell_range)
    anchor = sheet[cell_range.split(":", 1)[0]]
    anchor.value = text
    anchor.font = Font(bold=True)
    anchor.alignment = Alignment(wrap_text=True, vertical="top")
    if red_border:
        side = Side(style="medium", color="FF0000")
        anchor.border = Border(top=side, bottom=side, left=side, right=side)
    if shaded:
        anchor.fill = PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
        )


def _write_summary(
    sheet: Worksheet,
    calculation: dict[str, Any],
    roles: CitationRoles,
    generated_at: datetime,
    engine_version: str,
) -> None:
    sheet.title = "Summary"
    _banner(
        sheet,
        "A1:D1",
        "Excel Verification Pack - "
        f"{calculation['calculation_purpose']}",
    )
    rows: list[tuple[int, str, str]] = [
        (
            3,
            "Model",
            f"{calculation['model_id']} · v{calculation['model_version']}",
        ),
        (4, "Scenario", calculation["scenario"]["label"]),
        (5, "Test Date", calculation["test_date"]),
        (6, "Evaluation date", calculation["evaluation_date"]),
        (7, "Calculated LTV", calculation["outputs"]["ltv_display"]),
        (
            8,
            "Selected threshold",
            f"{calculation['selected_threshold']['percent']}% - "
            f"{roles.threshold['document_title']}, "
            f"{roles.threshold['locator']}",
        ),
        (9, "Headroom", calculation["outputs"]["headroom_display"]),
        (
            10,
            "Arithmetic status",
            _status_label(calculation["outputs"]["arithmetic_status"]),
        ),
    ]
    for row, label, value in rows:
        _set_text(sheet.cell(row=row, column=1), label)
        _set_text(sheet.cell(row=row, column=2), value)

    _banner(
        sheet,
        "A12:D14",
        _review_banner_text(calculation),
        red_border=True,
    )

    waiver_observation = calculation["waiver_observation"]
    if waiver_observation is not None:
        waiver_text = (
            "Limited waiver relief applies from "
            f"{waiver_observation['relief_above_percent']}% up to "
            f"{waiver_observation['relief_up_to_percent']}% for the "
            f"{waiver_observation['test_date']} Test Date only. "
            "This range is not the selected threshold."
        )
    else:
        waiver_text = "No waiver is relevant to this calculation."
    _banner(sheet, "A16:D18", waiver_text, shaded=True)

    _set_text(sheet.cell(row=20, column=1), "Provenance")
    _set_text(
        sheet.cell(row=20, column=2),
        f"engine {engine_version} · {calculation['model_id']} "
        f"v{calculation['model_version']} · scenario "
        f"{calculation['scenario']['scenario_id']} · generated "
        f"{generated_at.isoformat()}",
    )
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 60


def _write_inputs(
    sheet: Worksheet,
    calculation: dict[str, Any],
    roles: CitationRoles,
) -> None:
    headers = (
        "Field",
        "Canonical value (engine, exact)",
        "Recomputation value",
        "Currency",
        "Effective date",
        "Document title",
        "Locator",
        "Page",
        "Passage ID",
        "Exact supporting passage",
    )
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column)
        _set_text(cell, header)
        cell.font = Font(bold=True)

    debt_input = calculation["source_inputs"]["debt_amount"]
    valuation_input = calculation["source_inputs"]["valuation_amount"]
    rows: list[tuple[str, str, str | None, str, str, dict[str, Any]]] = [
        (
            "debt_amount",
            debt_input["value"],
            AMOUNT_FORMAT,
            calculation["currency"],
            debt_input["effective_date"],
            roles.debt,
        ),
        (
            "valuation_amount",
            valuation_input["value"],
            AMOUNT_FORMAT,
            calculation["currency"],
            valuation_input["effective_date"],
            roles.valuation,
        ),
        (
            "threshold_percent",
            calculation["selected_threshold"]["percent"],
            PERCENT_FORMAT,
            "",
            "",
            roles.threshold,
        ),
        (
            "ltv_definition",
            calculation["formula"]["expression"],
            None,
            "",
            "",
            roles.definition,
        ),
    ]
    for offset, (field, canonical, numeric_format, currency, effective,
                 citation) in enumerate(rows):
        row = 2 + offset
        _set_text(sheet.cell(row=row, column=1), field)
        _set_text(sheet.cell(row=row, column=2), canonical)
        if numeric_format is not None:
            numeric_cell = sheet.cell(row=row, column=3)
            numeric_cell.value = float(Decimal(canonical))
            numeric_cell.number_format = numeric_format
        _set_text(sheet.cell(row=row, column=4), currency)
        _set_text(sheet.cell(row=row, column=5), effective)
        _set_text(sheet.cell(row=row, column=6), citation["document_title"])
        _set_text(sheet.cell(row=row, column=7), citation["locator"])
        page_cell = sheet.cell(row=row, column=8)
        page_cell.value = int(citation["page"])
        page_cell.number_format = "0"
        _set_text(sheet.cell(row=row, column=9), citation["passage_id"])
        passage_cell = sheet.cell(row=row, column=10)
        _set_text(passage_cell, citation["supporting_passage"])
        passage_cell.alignment = Alignment(wrap_text=True, vertical="top")

    _banner(
        sheet,
        "A7:J9",
        f"{_review_banner_text(calculation)} {INPUTS_FOOTER}",
        red_border=True,
    )
    for column, width in zip("ABCDEFGHIJ", (18, 32, 20, 10, 14, 28, 22, 6,
                                            14, 60)):
        sheet.column_dimensions[column].width = width


def _write_calculation(sheet: Worksheet) -> None:
    _set_text(sheet.cell(row=1, column=1), "Metric")
    _set_text(sheet.cell(row=1, column=2), "Excel recomputation")
    sheet.cell(row=1, column=1).font = Font(bold=True)
    sheet.cell(row=1, column=2).font = Font(bold=True)
    rows = (
        (2, "LTV %", "=Inputs!C2/Inputs!C3*100", PERCENT_FORMAT),
        (3, "Headroom (pp)", "=Inputs!C4-B2", PERCENT_FORMAT),
        (
            4,
            "Maximum debt at threshold",
            "=Inputs!C3*(Inputs!C4/100)",
            AMOUNT_FORMAT,
        ),
        (5, "Debt capacity headroom", "=B4-Inputs!C2", AMOUNT_FORMAT),
        (
            6,
            "Minimum valuation at threshold",
            "=Inputs!C2/(Inputs!C4/100)",
            AMOUNT_FORMAT,
        ),
    )
    for row, label, formula, number_format in rows:
        _set_text(sheet.cell(row=row, column=1), label)
        formula_cell = sheet.cell(row=row, column=2)
        formula_cell.value = formula
        formula_cell.number_format = number_format
    _banner(sheet, "A8:B10", CALCULATION_FOOTER)
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 26


def _write_parity(sheet: Worksheet, calculation: dict[str, Any]) -> None:
    _banner(sheet, "A1:E1", PARITY_BANNER)
    headers = (
        "Metric",
        "Engine canonical value (authoritative)",
        "Excel recomputed value",
        "Delta",
        "Classification",
    )
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=column)
        _set_text(cell, header)
        cell.font = Font(bold=True)
    for offset, (metric, calculation_cell) in enumerate(_PARITY_METRICS):
        row = 3 + offset
        _set_text(sheet.cell(row=row, column=1), metric)
        _set_text(
            sheet.cell(row=row, column=2),
            calculation["outputs"][metric],
        )
        sheet.cell(row=row, column=3).value = (
            f"=Calculation!{calculation_cell}"
        )
        sheet.cell(row=row, column=4).value = f"=VALUE(B{row})-C{row}"
        sheet.cell(row=row, column=5).value = (
            f'=IF(D{row}=0,"Exact match",'
            '"Floating-point / precision artifact")'
        )
    for column, width in zip("ABCDE", (30, 44, 24, 18, 34)):
        sheet.column_dimensions[column].width = width


def _write_threshold_resolution(
    sheet: Worksheet,
    calculation: dict[str, Any],
    roles: CitationRoles,
) -> None:
    waiver_observation = calculation["waiver_observation"]
    threshold_source = (
        f"{roles.threshold['document_title']}, "
        f"{roles.threshold['locator']}, page {roles.threshold['page']}"
    )
    rows: list[tuple[int, str, str, bool]] = [
        (2, "Test Date", calculation["test_date"], False),
        (3, "Evaluation date", calculation["evaluation_date"], False),
        (
            4,
            "Amendment active",
            str(calculation["selected_threshold"]["amendment_active"]),
            False,
        ),
        (5, "Controlling document", threshold_source, False),
        (
            6,
            "Controlling passage (exact quote)",
            roles.threshold["supporting_passage"],
            False,
        ),
        (
            8,
            "Waiver relevant",
            str(waiver_observation is not None),
            False,
        ),
        (
            9,
            "Waiver test date",
            waiver_observation["test_date"] if waiver_observation else "",
            False,
        ),
        (
            10,
            "Waiver numeric range (not a threshold)",
            (
                f"{waiver_observation['relief_above_percent']}% - "
                f"{waiver_observation['relief_up_to_percent']}% "
                "(not a threshold)"
                if waiver_observation
                else ""
            ),
            True,
        ),
        (
            11,
            "Waiver passage (exact quote)",
            roles.waiver["supporting_passage"] if roles.waiver else "",
            False,
        ),
        (
            12,
            "Does not amend threshold",
            (
                str(waiver_observation["does_not_amend_threshold"])
                if waiver_observation
                else ""
            ),
            False,
        ),
    ]
    for row, label, value, shaded in rows:
        _set_text(sheet.cell(row=row, column=1), label)
        value_cell = sheet.cell(row=row, column=2)
        _set_text(value_cell, value)
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")
        if shaded:
            fill = PatternFill(
                start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
            )
            sheet.cell(row=row, column=1).fill = fill
            value_cell.fill = fill
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 70


def _write_scenario(sheet: Worksheet, calculation: dict[str, Any]) -> None:
    scenario = calculation["scenario"]
    _set_text(sheet.cell(row=2, column=1), "Scenario label")
    _set_text(sheet.cell(row=2, column=2), scenario["label"])
    _set_text(sheet.cell(row=3, column=1), "Rationale")
    _set_text(sheet.cell(row=3, column=2), scenario["rationale"])
    row = 4
    _set_text(
        sheet.cell(row=row, column=1),
        "Assumptions (as stated, not source facts)",
    )
    assumptions = calculation["assumptions"] or [""]
    for assumption in assumptions:
        _set_text(
            sheet.cell(row=row, column=2),
            f"Assumption: {assumption}" if assumption else "",
        )
        row += 1
    row += 1
    _set_text(sheet.cell(row=row, column=1), "Debt used in this scenario")
    _set_text(
        sheet.cell(row=row, column=2),
        calculation["calculation_inputs"]["debt_amount"],
    )
    row += 1
    _set_text(
        sheet.cell(row=row, column=1), "Valuation used in this scenario"
    )
    _set_text(
        sheet.cell(row=row, column=2),
        calculation["calculation_inputs"]["valuation_amount"],
    )
    row += 2
    _banner(
        sheet,
        f"A{row}:B{row + 2}",
        "Scenario values are hypothetical assumptions layered on the "
        "sourced inputs shown on the Inputs sheet. They are never presented "
        "as source facts.",
        shaded=True,
    )
    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 70


def _normalize_zip(data: bytes, stamp: datetime) -> bytes:
    """Rewrite zip members with a fixed timestamp for deterministic bytes."""
    date_time = (
        max(stamp.year, 1980),
        stamp.month,
        stamp.day,
        stamp.hour,
        stamp.minute,
        stamp.second,
    )
    source = zipfile.ZipFile(io.BytesIO(data), "r")
    buffer = io.BytesIO()
    with zipfile.ZipFile(
        buffer, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6
    ) as target:
        for item in source.infolist():
            info = zipfile.ZipInfo(item.filename, date_time=date_time)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = item.external_attr
            target.writestr(info, source.read(item.filename))
    source.close()
    return buffer.getvalue()


def generate_verification_workbook(
    calculation: dict[str, Any],
    *,
    generated_at: datetime | None = None,
    engine_version: str = "1.3.0",
) -> tuple[bytes, dict[str, Any]]:
    """Build the six-sheet .xlsx verification pack for one calculated LTV
    result.

    The calculation argument is the exact dict shape produced by
    AgreementIntelligenceEngine.calculate_ltv(). Abstaining payloads fail
    closed: no workbook, no partial bytes. The engine's decimal values are
    authoritative; the workbook's live formulas are a reviewer cross-check.
    """
    if not isinstance(calculation, dict):
        raise ValueError("calculation must be a JSON object.")
    status = calculation.get("status")
    if status != "calculated_human_review_required":
        raise WorkbookGenerationAbstained(
            "Workbook generation requires a calculated result; received "
            f"status {status!r}.",
            list(calculation.get("missing_information", [])),
        )
    missing_keys = _REQUIRED_KEYS - set(calculation)
    if missing_keys:
        raise ValueError(
            f"calculation is missing required keys: {sorted(missing_keys)}"
        )

    roles = _resolve_citation_roles(calculation)
    if generated_at is None:
        generated_at = datetime.now(timezone.utc)
    if generated_at.tzinfo is None:
        generated_at = generated_at.replace(tzinfo=timezone.utc)
    stamp = generated_at.astimezone(timezone.utc).replace(tzinfo=None)

    workbook = Workbook()
    workbook.properties.created = stamp
    workbook.properties.modified = stamp
    _write_summary(
        workbook.active, calculation, roles, generated_at, engine_version
    )
    _write_inputs(
        workbook.create_sheet("Inputs"), calculation, roles
    )
    _write_calculation(workbook.create_sheet("Calculation"))
    _write_parity(workbook.create_sheet("Parity"), calculation)
    _write_threshold_resolution(
        workbook.create_sheet("Threshold resolution"), calculation, roles
    )
    if calculation["scenario"]["scenario_id"] != "baseline":
        _write_scenario(workbook.create_sheet("Scenario"), calculation)

    raw = io.BytesIO()
    workbook.save(raw)
    workbook_bytes = _normalize_zip(raw.getvalue(), stamp)
    provenance = {
        "engine_version": engine_version,
        "model_id": calculation["model_id"],
        "model_version": int(calculation["model_version"]),
        "scenario_id": calculation["scenario"]["scenario_id"],
        "generated_at": generated_at.isoformat(),
        "sha256": hashlib.sha256(workbook_bytes).hexdigest(),
    }
    return workbook_bytes, provenance
