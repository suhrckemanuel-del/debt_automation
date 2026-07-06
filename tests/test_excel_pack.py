from __future__ import annotations

import copy
import hashlib
import io
import unittest
import zipfile
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from agreement_intelligence.engine import AgreementIntelligenceEngine
from agreement_intelligence.excel_pack import (
    SUMMARY_PROVENANCE_CELL,
    WorkbookGenerationAbstained,
    generate_verification_workbook,
)

FIXED_TS = datetime(2026, 7, 5, 12, 0, 0, tzinfo=timezone.utc)
LATER_TS = datetime(2026, 7, 5, 18, 30, 0, tzinfo=timezone.utc)


def _load(workbook_bytes: bytes):
    return load_workbook(io.BytesIO(workbook_bytes), data_only=False)


class ExcelPackTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        root = Path(__file__).resolve().parents[1]
        cls.engine = AgreementIntelligenceEngine(root)
        cls.calculation = cls.engine.calculate_ltv()

    def test_same_fixed_timestamp_is_byte_identical(self) -> None:
        first, first_provenance = generate_verification_workbook(
            self.calculation, generated_at=FIXED_TS
        )
        second, second_provenance = generate_verification_workbook(
            self.calculation, generated_at=FIXED_TS
        )
        self.assertEqual(
            hashlib.sha256(first).hexdigest(),
            hashlib.sha256(second).hexdigest(),
        )
        self.assertEqual(first_provenance, second_provenance)
        self.assertEqual(
            first_provenance["sha256"],
            hashlib.sha256(first).hexdigest(),
        )

    def test_varying_timestamp_only_changes_timestamp_cell(self) -> None:
        first, _ = generate_verification_workbook(
            self.calculation, generated_at=FIXED_TS
        )
        second, _ = generate_verification_workbook(
            self.calculation, generated_at=LATER_TS
        )
        first_book = _load(first)
        second_book = _load(second)
        self.assertEqual(first_book.sheetnames, second_book.sheetnames)
        for sheet_name in first_book.sheetnames:
            first_sheet = first_book[sheet_name]
            second_sheet = second_book[sheet_name]
            for first_row, second_row in zip(
                first_sheet.iter_rows(), second_sheet.iter_rows()
            ):
                for first_cell, second_cell in zip(first_row, second_row):
                    if (
                        sheet_name == "Summary"
                        and first_cell.coordinate == SUMMARY_PROVENANCE_CELL
                    ):
                        self.assertNotEqual(
                            first_cell.value, second_cell.value
                        )
                        continue
                    self.assertEqual(
                        first_cell.value,
                        second_cell.value,
                        f"{sheet_name}!{first_cell.coordinate} differs",
                    )

    def test_parity_delta_zero_for_baseline_metrics(self) -> None:
        workbook_bytes, _ = generate_verification_workbook(
            self.calculation, generated_at=FIXED_TS
        )
        parity = _load(workbook_bytes)["Parity"]
        engine_values = {
            parity.cell(row=row, column=1).value: parity.cell(
                row=row, column=2
            ).value
            for row in range(3, 8)
        }
        debt = 71_000_000.0
        valuation = 100_000_000.0
        threshold = 70.0
        float_recomputation = {
            "ltv_percent": debt / valuation * 100,
            "headroom_percentage_points": threshold - debt / valuation * 100,
            "maximum_debt_at_threshold": valuation * (threshold / 100),
            "debt_capacity_headroom": valuation * (threshold / 100) - debt,
        }
        for metric, recomputed in float_recomputation.items():
            self.assertEqual(
                Decimal(str(recomputed)).normalize(),
                Decimal(engine_values[metric]).normalize(),
                f"{metric} float recomputation must equal the engine value",
            )
        for row in range(3, 8):
            self.assertEqual(
                parity.cell(row=row, column=4).value,
                f"=VALUE(B{row})-C{row}",
            )
            self.assertIn(
                "Floating-point / precision artifact",
                parity.cell(row=row, column=5).value,
            )

    def test_parity_minimum_valuation_is_labelled_artifact(self) -> None:
        workbook_bytes, _ = generate_verification_workbook(
            self.calculation, generated_at=FIXED_TS
        )
        parity = _load(workbook_bytes)["Parity"]
        row = next(
            row
            for row in range(3, 8)
            if parity.cell(row=row, column=1).value
            == "minimum_valuation_at_threshold"
        )
        engine_value = Decimal(parity.cell(row=row, column=2).value)
        float_value = Decimal(str(71_000_000.0 / (70.0 / 100)))
        self.assertNotEqual(
            engine_value.normalize(),
            float_value.normalize(),
            "The repeating decimal must not equal any float recomputation",
        )
        banner = parity.cell(row=1, column=1).value
        self.assertIn("authoritative", banner)

    def test_workbook_has_no_macro_or_external_link_parts(self) -> None:
        workbook_bytes, _ = generate_verification_workbook(
            self.calculation, generated_at=FIXED_TS
        )
        with zipfile.ZipFile(io.BytesIO(workbook_bytes)) as archive:
            names = archive.namelist()
        self.assertNotIn("xl/vbaProject.bin", names)
        self.assertFalse(
            [name for name in names if name.startswith("xl/externalLinks/")]
        )
        reloaded = _load(workbook_bytes)
        self.assertIsNone(reloaded.vba_archive)

    def test_inputs_sheet_citations_match_payload(self) -> None:
        workbook_bytes, _ = generate_verification_workbook(
            self.calculation, generated_at=FIXED_TS
        )
        inputs = _load(workbook_bytes)["Inputs"]
        sources = {
            (source["document_id"], source["locator"]): source
            for source in self.calculation["sources"]
        }
        source_inputs = self.calculation["source_inputs"]
        threshold = self.calculation["selected_threshold"]
        expected_rows = {
            "debt_amount": (
                source_inputs["debt_amount"]["value"],
                sources[
                    (
                        source_inputs["debt_amount"]["document_id"],
                        source_inputs["debt_amount"]["locator"],
                    )
                ],
            ),
            "valuation_amount": (
                source_inputs["valuation_amount"]["value"],
                sources[
                    (
                        source_inputs["valuation_amount"]["document_id"],
                        source_inputs["valuation_amount"]["locator"],
                    )
                ],
            ),
            "threshold_percent": (
                threshold["percent"],
                sources[(threshold["document_id"], threshold["locator"])],
            ),
        }
        for row in range(2, 5):
            field = inputs.cell(row=row, column=1).value
            canonical, citation = expected_rows[field]
            self.assertEqual(inputs.cell(row=row, column=2).value, canonical)
            self.assertEqual(
                inputs.cell(row=row, column=6).value,
                citation["document_title"],
            )
            self.assertEqual(
                inputs.cell(row=row, column=7).value, citation["locator"]
            )
            self.assertEqual(
                inputs.cell(row=row, column=8).value, citation["page"]
            )
            self.assertEqual(
                inputs.cell(row=row, column=9).value,
                citation["passage_id"],
            )
            self.assertEqual(
                inputs.cell(row=row, column=10).value,
                citation["supporting_passage"],
            )
        self.assertEqual(
            inputs.cell(row=5, column=1).value, "ltv_definition"
        )
        self.assertEqual(
            inputs.cell(row=5, column=2).value,
            self.calculation["formula"]["expression"],
        )

    def test_waiver_ceiling_never_in_threshold_cell(self) -> None:
        waiver = self.calculation["waiver_observation"]
        self.assertEqual(
            Decimal(waiver["relief_up_to_percent"]), Decimal("72")
        )
        workbook_bytes, _ = generate_verification_workbook(
            self.calculation, generated_at=FIXED_TS
        )
        book = _load(workbook_bytes)
        for sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            for row in sheet.iter_rows():
                label = row[0].value
                if not isinstance(label, str):
                    continue
                if "threshold" not in label.lower():
                    continue
                if "not a threshold" in label.lower():
                    continue
                for cell in row[1:]:
                    value = cell.value
                    if not isinstance(value, str):
                        continue
                    self.assertNotEqual(value, "72")
                    self.assertFalse(
                        value.startswith("72.") or "72.00%" in value
                        or value.startswith("72%"),
                        f"Waiver ceiling in threshold-labelled cell "
                        f"{sheet_name}!{cell.coordinate}: {value!r}",
                    )

    def test_calculation_unavailable_payload_raises_typed_abstention(
        self,
    ) -> None:
        unavailable = {
            "model_id": "ltv-v1",
            "status": "calculation_unavailable",
            "missing_information": ["Reviewed financial model definition."],
            "human_review_required": True,
            "sources": [],
        }
        with self.assertRaises(WorkbookGenerationAbstained) as context:
            generate_verification_workbook(
                unavailable, generated_at=FIXED_TS
            )
        self.assertEqual(
            context.exception.missing_information,
            ["Reviewed financial model definition."],
        )

    def test_scenario_sheet_only_for_non_baseline(self) -> None:
        baseline_bytes, baseline_provenance = (
            generate_verification_workbook(
                self.calculation, generated_at=FIXED_TS
            )
        )
        self.assertNotIn("Scenario", _load(baseline_bytes).sheetnames)
        self.assertEqual(baseline_provenance["scenario_id"], "baseline")

        scenario_calculation = self.engine.calculate_ltv(
            scenario_id="valuation-down-10"
        )
        scenario_bytes, scenario_provenance = (
            generate_verification_workbook(
                scenario_calculation, generated_at=FIXED_TS
            )
        )
        book = _load(scenario_bytes)
        self.assertIn("Scenario", book.sheetnames)
        self.assertEqual(
            scenario_provenance["scenario_id"],
            scenario_calculation["scenario"]["scenario_id"],
        )
        scenario_sheet = book["Scenario"]
        values = [
            cell.value
            for row in scenario_sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str)
        ]
        self.assertTrue(
            any("not source facts" in value for value in values)
        )

    def test_tampered_source_set_fails_closed(self) -> None:
        tampered = copy.deepcopy(self.calculation)
        tampered["sources"] = tampered["sources"][:2]
        with self.assertRaises(WorkbookGenerationAbstained):
            generate_verification_workbook(tampered, generated_at=FIXED_TS)


if __name__ == "__main__":
    unittest.main()
