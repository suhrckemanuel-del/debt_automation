from __future__ import annotations

import base64
import copy
import hashlib
import io
import json
import threading
import unittest
from http.server import ThreadingHTTPServer
from pathlib import Path
from urllib.error import HTTPError
from urllib.request import Request, urlopen

from openpyxl import load_workbook

from agreement_intelligence.api import make_engine_api_handler


class EngineApiContractTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.root = Path(__file__).resolve().parents[1]
        cls.server = ThreadingHTTPServer(
            ("127.0.0.1", 0),
            make_engine_api_handler(cls.root),
        )
        cls.thread = threading.Thread(
            target=cls.server.serve_forever,
            daemon=True,
        )
        cls.thread.start()
        cls.base_url = f"http://127.0.0.1:{cls.server.server_port}"

    @classmethod
    def tearDownClass(cls) -> None:
        cls.server.shutdown()
        cls.server.server_close()
        cls.thread.join(timeout=5)

    def test_health_exposes_contract_version(self) -> None:
        with urlopen(f"{self.base_url}/health", timeout=5) as response:
            payload = json.loads(response.read())
        self.assertEqual(
            payload,
            {"status": "ok", "contract_version": "1.3.0"},
        )

    def test_position_exposes_resolved_facts_and_exact_sources(self) -> None:
        with urlopen(
            f"{self.base_url}/v1/workspaces/demo/position"
            "?as_of=2026-07-02",
            timeout=5,
        ) as response:
            payload = json.loads(response.read())

        self.assertEqual(payload["support_status"], "supported")
        self.assertEqual(payload["ltv"]["current_threshold"], 70.0)
        self.assertTrue(payload["ltv"]["amendment_active"])
        self.assertTrue(payload["ltv"]["waiver_relevant"])
        self.assertTrue(payload["ltv"]["distribution_condition_active"])
        self.assertEqual(len(payload["sources"]), 4)
        self.assertTrue(
            all(source["supporting_passage"] for source in payload["sources"])
        )

    def test_position_resolves_each_review_date_without_ui_rules(self) -> None:
        cases = {
            "2026-04-01": (65.0, False, False, False),
            "2026-07-02": (70.0, True, True, True),
            "2026-10-01": (70.0, True, False, False),
            "2027-01-02": (65.0, False, False, False),
        }
        for as_of, expected in cases.items():
            with self.subTest(as_of=as_of):
                with urlopen(
                    f"{self.base_url}/v1/workspaces/demo/position"
                    f"?as_of={as_of}",
                    timeout=5,
                ) as response:
                    payload = json.loads(response.read())
                ltv = payload["ltv"]
                self.assertEqual(
                    (
                        ltv["current_threshold"],
                        ltv["amendment_active"],
                        ltv["waiver_relevant"],
                        ltv["distribution_condition_active"],
                    ),
                    expected,
                )

    def test_answer_preserves_structured_citations_and_abstention(self) -> None:
        body = json.dumps(
            {
                "question": "Does the facility require a debt yield covenant?",
                "as_of": "2026-07-02",
            }
        ).encode()
        request = Request(
            f"{self.base_url}/v1/workspaces/demo/answers",
            data=body,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urlopen(request, timeout=5) as response:
            payload = json.loads(response.read())
        self.assertEqual(payload["support_status"], "source_not_found")
        self.assertEqual(payload["short_answer"], "Source not found.")
        self.assertEqual(payload["sources"], [])

    def test_ltv_calculation_contract_returns_trace_and_sources(self) -> None:
        body = json.dumps(
            {"model_id": "ltv-v1", "scenario_id": None}
        ).encode()
        request = Request(
            f"{self.base_url}/v1/workspaces/demo/models/ltv-calculations",
            data=body,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urlopen(request, timeout=5) as response:
            payload = json.loads(response.read())
        self.assertEqual(payload["outputs"]["ltv_percent"], "71")
        self.assertEqual(
            payload["outputs"]["arithmetic_status"],
            "above_selected_threshold",
        )
        self.assertEqual(payload["selected_threshold"]["percent"], "70")
        self.assertEqual(len(payload["sources"]), 5)
        self.assertTrue(payload["human_review_required"])

    def _calculated_ltv(self) -> dict:
        body = json.dumps(
            {"model_id": "ltv-v1", "scenario_id": None}
        ).encode()
        request = Request(
            f"{self.base_url}/v1/workspaces/demo/models/ltv-calculations",
            data=body,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urlopen(request, timeout=5) as response:
            return json.loads(response.read())

    def _post_workbook(self, payload: dict) -> tuple[int, dict]:
        request = Request(
            f"{self.base_url}/v1/workspaces/demo/models/ltv-calculations"
            "/verification-workbook",
            data=json.dumps(payload).encode(),
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        try:
            with urlopen(request, timeout=10) as response:
                return response.status, json.loads(response.read())
        except HTTPError as error:
            return error.code, json.loads(error.read())

    def test_verification_workbook_round_trip(self) -> None:
        status, payload = self._post_workbook(self._calculated_ltv())
        self.assertEqual(status, 200)
        workbook_bytes = base64.b64decode(payload["workbook_base64"])
        self.assertEqual(
            payload["provenance"]["sha256"],
            hashlib.sha256(workbook_bytes).hexdigest(),
        )
        self.assertEqual(payload["provenance"]["scenario_id"], "baseline")
        book = load_workbook(io.BytesIO(workbook_bytes))
        self.assertEqual(
            book.sheetnames,
            [
                "Summary",
                "Inputs",
                "Calculation",
                "Parity",
                "Threshold resolution",
            ],
        )
        self.assertEqual(book["Summary"]["B7"].value, "71.00%")

    def test_verification_workbook_endpoint_rejects_abstaining_payload(
        self,
    ) -> None:
        status, payload = self._post_workbook(
            {
                "model_id": "ltv-v1",
                "status": "calculation_unavailable",
                "missing_information": [
                    "Reviewed financial model definition."
                ],
                "human_review_required": True,
                "sources": [],
            }
        )
        self.assertEqual(status, 400)
        self.assertIn("refused", payload["error"])
        self.assertIn(
            "Reviewed financial model definition.", payload["error"]
        )
        self.assertNotIn("workbook_base64", payload)

    def test_verification_workbook_endpoint_rejects_tampered_citation(
        self,
    ) -> None:
        tampered = copy.deepcopy(self._calculated_ltv())
        tampered["sources"][0]["supporting_passage"] = (
            "This text does not appear in the corpus."
        )
        status, payload = self._post_workbook(tampered)
        self.assertEqual(status, 400)
        self.assertIn("does not match the live corpus", payload["error"])
        self.assertNotIn("workbook_base64", payload)

    def test_verification_workbook_rejects_waiver_relabelled_as_threshold(
        self,
    ) -> None:
        tampered = copy.deepcopy(self._calculated_ltv())
        tampered["selected_threshold"]["percent"] = (
            tampered["waiver_observation"]["relief_up_to_percent"]
        )
        status, payload = self._post_workbook(tampered)
        self.assertEqual(status, 400)
        self.assertIn(
            "does not match the resolved contractual position",
            payload["error"],
        )
        self.assertNotIn("workbook_base64", payload)


if __name__ == "__main__":
    unittest.main()
